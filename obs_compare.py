"""
OBS Gap Analysis: G-TRN A08001 Standard PTO vs Existing Project OBS
Generates a colour-coded Excel workbook with:
  - Sheet 1 : Gap Analysis (one row per standard role, one column per project)
  - Sheet 2 : R152B Detail (all existing roles mapped back to standard)
  - Sheet 3 : Legend & Instructions

Usage:
  python obs_compare.py
Output:
  2_Output/OBS_Gap_Analysis.xlsx
"""

import os
import sys
import re
import pypdf
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule
from rapidfuzz import fuzz, process

sys.stdout.reconfigure(encoding="utf-8")

# ---------------------------------------------------------------------------
# 1. STANDARD ROLES from G-TRN A08001 (Project Execution Team)
# ---------------------------------------------------------------------------
STANDARD_ROLES = [
    # --- Core roles (SRS + VHC) ---
    {"role": "Project Manager",                    "abbr": "PM",   "lob": "SRS+VHC", "mandatory": True},
    {"role": "Project Controller",                 "abbr": "PJC",  "lob": "SRS+VHC", "mandatory": True},
    {"role": "Project Planner",                    "abbr": "PJP",  "lob": "SRS+VHC", "mandatory": True},
    {"role": "Project Engineer",                   "abbr": "PE",   "lob": "SRS+VHC", "mandatory": True},
    {"role": "Contract Manager",                   "abbr": "CM",   "lob": "SRS+VHC", "mandatory": True},
    {"role": "Project Quality Manager",            "abbr": "PQM",  "lob": "SRS+VHC", "mandatory": True},
    {"role": "HSE Referent / HSE Manager",         "abbr": "HSE",  "lob": "SRS+VHC", "mandatory": True},
    # --- SRS-specific roles ---
    {"role": "Project RAMS Manager",               "abbr": "PRAMS","lob": "SRS",     "mandatory": False},
    {"role": "Supply Chain & Procurement Manager", "abbr": "SCPM", "lob": "SRS",     "mandatory": False},
    {"role": "Construction / Installation Manager","abbr": "CONSTR","lob": "SRS",    "mandatory": False},
    {"role": "Commissioning / T&C Manager",        "abbr": "COMM", "lob": "SRS",     "mandatory": False},
    {"role": "Warranty Manager",                   "abbr": "WM",   "lob": "SRS",     "mandatory": False},
    # --- VHC-specific roles (included for completeness) ---
    {"role": "Project RAM Manager",                "abbr": "RAM",  "lob": "VHC",     "mandatory": False},
    {"role": "Project Safety Manager",             "abbr": "PSM",  "lob": "VHC",     "mandatory": False},
    {"role": "Project Procurement Manager",        "abbr": "PPM",  "lob": "VHC",     "mandatory": False},
    {"role": "Test Manager",                       "abbr": "TM",   "lob": "VHC",     "mandatory": False},
    {"role": "Project Operations Manager",         "abbr": "POM",  "lob": "VHC",     "mandatory": False},
]

# Keywords used for fuzzy matching each standard role to existing OBS role titles
ROLE_KEYWORDS = {
    "Project Manager":                    ["project manager", "pm", "program manager", "programme manager"],
    "Project Controller":                 ["controller", "financial controller", "project controller", "finance"],
    "Project Planner":                    ["planner", "planning", "scheduler"],
    "Project Engineer":                   ["project engineer", "system engineering", "design authority", "pda", "pe"],
    "Contract Manager":                   ["contract manager", "contracts"],
    "Project Quality Manager":            ["quality", "qam", "qa manager", "quality assurance", "qm"],
    "HSE Referent / HSE Manager":         ["hse", "health safety", "safety manager", "wsho", "environmental"],
    "Project RAMS Manager":               ["rams", "ram manager", "rams manager"],
    "Supply Chain & Procurement Manager": ["supply chain", "procurement", "scpm", "sourcing", "pscl"],
    "Construction / Installation Manager":["construction", "installation", "deployment"],
    "Commissioning / T&C Manager":        ["commissioning", "t&c", "test and commissioning", "ivvq", "ival"],
    "Warranty Manager":                   ["warranty"],
    "Project RAM Manager":                ["ram specialist", "ram manager", "reliability"],
    "Project Safety Manager":             ["safety manager", "project safety", "system assurance", "safety architect"],
    "Project Procurement Manager":        ["procurement manager", "ppm"],
    "Test Manager":                       ["test manager"],
    "Project Operations Manager":         ["operations manager"],
}

# ---------------------------------------------------------------------------
# 2. EXISTING PROJECT DATA
# ---------------------------------------------------------------------------
def extract_obs_roles_from_pdf(pdf_path: str) -> list[dict]:
    """Extract unique role titles from an OBS PDF (text extraction)."""
    reader = pypdf.PdfReader(pdf_path)
    all_text = ""
    for page in reader.pages:
        t = page.extract_text() or ""
        all_text += t + "\n"

    # Deduplicate lines, clean up repeated text (PDF charts often duplicate)
    lines = all_text.split("\n")
    seen = set()
    roles = []
    noise = {
        "hitachi rail", "refer to the sap", "other roles omitted",
        "organization breakdown", "r152b sa1", "project organization",
    }
    for line in lines:
        line = line.strip()
        if len(line) < 4:
            continue
        # Skip lines that are just person names (heuristic: contains only name chars)
        if re.match(r'^[A-Z][a-z]+(\s+[A-Z][a-z]+){1,4}$', line):
            continue
        # Skip noise
        if any(n in line.lower() for n in noise):
            continue
        # Skip lines that look like entity headers
        if line.lower() in ["hitachi rail france", "hitachi rail singapore",
                             "hitachi rail canada"]:
            continue
        key = line.lower().strip()
        if key not in seen and len(key) > 3:
            seen.add(key)
            roles.append({"title": line, "entity": "Various"})
    return roles


def manual_r152b_roles() -> list[dict]:
    """Manually curated role list from R152B SA1 OBS, de-duplicated, with entity."""
    return [
        # ---- Hitachi Rail Singapore (lead entity) ----
        {"title": "Singapore PM",                        "entity": "HR Singapore"},
        {"title": "Singapore Deputy PM",                 "entity": "HR Singapore"},
        {"title": "PMO",                                 "entity": "HR Singapore"},
        {"title": "Financial Controller",                "entity": "HR Singapore"},
        {"title": "Project Design Authority (PDA)",      "entity": "HR Singapore"},
        {"title": "Interface Manager",                   "entity": "HR Singapore"},
        {"title": "Configuration Manager",               "entity": "HR Singapore"},
        {"title": "Contract Manager",                    "entity": "HR Singapore"},
        {"title": "Project Admin",                       "entity": "HR Singapore"},
        {"title": "Document Controller",                 "entity": "HR Singapore"},
        {"title": "Planner",                             "entity": "HR Singapore"},
        {"title": "System Engineering Manager",          "entity": "HR Singapore"},
        {"title": "HW Design Lead",                      "entity": "HR Singapore"},
        {"title": "Deployment Manager",                  "entity": "HR Singapore"},
        {"title": "T&C Manager (Test & Commissioning)",  "entity": "HR Singapore"},
        {"title": "Supply Chain Manager",                "entity": "HR Singapore"},
        {"title": "Senior Supply Chain Manager",         "entity": "HR Singapore"},
        {"title": "Supply Chain Executive",              "entity": "HR Singapore"},
        {"title": "Project Procurement Manager",         "entity": "HR Singapore"},
        {"title": "PO Admin",                            "entity": "HR Singapore"},
        {"title": "Sourcing & Contracting",              "entity": "HR Singapore"},
        {"title": "Supplier Performance",                "entity": "HR Singapore"},
        {"title": "DLP Manager",                         "entity": "HR Singapore"},
        {"title": "DLP Engineer",                        "entity": "HR Singapore"},
        {"title": "Installation Lead",                   "entity": "HR Singapore"},
        {"title": "HSE Manager",                         "entity": "HR Singapore"},
        {"title": "WSHO Officer",                        "entity": "HR Singapore"},
        {"title": "Environmental Control Officer",       "entity": "HR Singapore"},
        {"title": "Technical Safety Specialist",         "entity": "HR Singapore"},
        {"title": "Local System Assurance Manager",      "entity": "HR Singapore"},
        {"title": "DCS Lead",                            "entity": "HR Singapore"},
        {"title": "Software Engineers",                  "entity": "HR Singapore"},
        {"title": "Warranty",                            "entity": "HR Singapore"},
        {"title": "RAM Specialist",                      "entity": "HR Singapore"},
        {"title": "IVAL Specialist (IV&V)",              "entity": "HR Singapore"},
        {"title": "Safety Engineering Specialists",      "entity": "HR Singapore"},
        {"title": "Chief Safety Architect",              "entity": "HR Singapore"},
        {"title": "Safety Assurance Specialist",         "entity": "HR Singapore"},
        {"title": "Quality Assurance Manager",           "entity": "HR Singapore"},
        # ---- Hitachi Rail France ----
        {"title": "France PM",                           "entity": "HR France"},
        {"title": "PMO",                                 "entity": "HR France"},
        {"title": "Financial Controller",                "entity": "HR France"},
        {"title": "System Engineering & Software Manager","entity": "HR France"},
        {"title": "Configuration & Documentation Manager","entity": "HR France"},
        {"title": "PDA (Project Design Authority)",      "entity": "HR France"},
        {"title": "Quality Assurance Manager",           "entity": "HR France"},
        {"title": "PPM (Project Programme Manager)",     "entity": "HR France"},
        {"title": "RAMS Manager",                        "entity": "HR France"},
        {"title": "Technical Safety Specialist",         "entity": "HR France"},
        # ---- Hitachi Rail Canada ----
        {"title": "Canada PM",                           "entity": "HR Canada"},
        {"title": "Financial Controller",                "entity": "HR Canada"},
        # ---- Cross-entity / Programme level ----
        {"title": "PPM/PSCL (Programme & Supply Chain Lead)", "entity": "Programme"},
        {"title": "System Assurance Manager",            "entity": "Programme"},
        {"title": "System Engineering Manager",          "entity": "Programme"},
        {"title": "Deployment Manager",                  "entity": "Programme"},
        {"title": "Project Safety Manager",              "entity": "Programme"},
        {"title": "IVVQ Manager",                        "entity": "Programme"},
        {"title": "Project Design Authority",            "entity": "Programme"},
        {"title": "Quality Assurance Manager",           "entity": "Programme"},
        {"title": "Planner",                             "entity": "Programme"},
        {"title": "RAMS Manager",                        "entity": "Programme"},
    ]


# ---------------------------------------------------------------------------
# 3. MATCHING LOGIC
# ---------------------------------------------------------------------------
STATUS_MATCHED   = "MATCHED"
STATUS_PARTIAL   = "PARTIAL MATCH"
STATUS_MISSING   = "MISSING"
STATUS_EXTRA     = "EXTRA (no G-TRN equivalent)"

def match_role(standard_role: str, obs_roles: list[dict]) -> dict:
    """
    Try to match a standard role to the list of OBS role titles.
    Returns dict: {status, matched_title, entity, score}
    """
    keywords = ROLE_KEYWORDS.get(standard_role, [standard_role.lower()])
    obs_titles = [r["title"] for r in obs_roles]

    best_score = 0
    best_title = None
    best_entity = None

    for obs in obs_roles:
        title_lower = obs["title"].lower()
        for kw in keywords:
            # exact substring
            if kw in title_lower:
                score = 100
            else:
                score = fuzz.partial_ratio(kw, title_lower)
            if score > best_score:
                best_score = score
                best_title = obs["title"]
                best_entity = obs["entity"]

    if best_score >= 90:
        return {"status": STATUS_MATCHED, "matched_title": best_title,
                "entity": best_entity, "score": best_score}
    elif best_score >= 65:
        return {"status": STATUS_PARTIAL, "matched_title": best_title,
                "entity": best_entity, "score": best_score}
    else:
        return {"status": STATUS_MISSING, "matched_title": "",
                "entity": "", "score": best_score}


def find_extra_roles(obs_roles: list[dict], matched_titles: set) -> list[dict]:
    """Roles in the project OBS that were not matched to any standard role."""
    extras = []
    for r in obs_roles:
        if r["title"] not in matched_titles:
            extras.append(r)
    return extras


# ---------------------------------------------------------------------------
# 4. EXCEL REPORT BUILDER
# ---------------------------------------------------------------------------
# Colour palette
CLR_GREEN  = "C6EFCE"   # matched
CLR_ORANGE = "FFEB9C"   # partial
CLR_RED    = "FFC7CE"   # missing
CLR_BLUE   = "BDD7EE"   # extra
CLR_HEADER = "1F4E79"   # dark blue header
CLR_SUBHDR = "2E75B6"   # section header
CLR_MANDATORY_BG = "FFF2CC"  # mandatory role row highlight

FILL_GREEN  = PatternFill("solid", fgColor=CLR_GREEN)
FILL_ORANGE = PatternFill("solid", fgColor=CLR_ORANGE)
FILL_RED    = PatternFill("solid", fgColor=CLR_RED)
FILL_BLUE   = PatternFill("solid", fgColor=CLR_BLUE)
FILL_HEADER = PatternFill("solid", fgColor=CLR_HEADER)
FILL_SUBHDR = PatternFill("solid", fgColor=CLR_SUBHDR)
FILL_MAND   = PatternFill("solid", fgColor=CLR_MANDATORY_BG)

FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
FONT_BOLD       = Font(bold=True, name="Calibri", size=10)
FONT_NORMAL     = Font(name="Calibri", size=10)
FONT_ITALIC     = Font(italic=True, name="Calibri", size=9, color="595959")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def status_fill(status: str) -> PatternFill:
    return {
        STATUS_MATCHED: FILL_GREEN,
        STATUS_PARTIAL: FILL_ORANGE,
        STATUS_MISSING: FILL_RED,
        STATUS_EXTRA:   FILL_BLUE,
    }.get(status, PatternFill())


def status_icon(status: str) -> str:
    return {
        STATUS_MATCHED: "✔ Matched",
        STATUS_PARTIAL: "~ Partial",
        STATUS_MISSING: "✘ MISSING",
        STATUS_EXTRA:   "+ Extra",
    }.get(status, status)


def apply_cell(ws, row, col, value, fill=None, font=None,
               alignment=None, border=True, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:
        c.fill = fill
    if font:
        c.font = font
    else:
        c.font = FONT_NORMAL
    if alignment:
        c.alignment = alignment
    if border:
        c.border = THIN_BORDER
    if number_format:
        c.number_format = number_format
    return c


# ---------------------------------------------------------------------------
# 5.  SHEET 1 – GAP ANALYSIS MATRIX
# ---------------------------------------------------------------------------
def build_gap_matrix(wb: Workbook, projects: list[dict]) -> None:
    ws = wb.active
    ws.title = "OBS Gap Analysis"
    ws.sheet_view.showGridLines = False

    # ---- Title block ----
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1,
                value="OBS GAP ANALYSIS  |  G-TRN A08001 Standard PTO vs Existing Project OBS")
    c.fill = FILL_HEADER
    c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
    c.alignment = CENTER

    ws.merge_cells("A2:H2")
    c = ws.cell(row=2, column=1,
                value="Green = Matched   |   Orange = Partial / Different Name   |"
                      "   Red = MISSING (GAP)   |   Blue = Extra role (not in G-TRN standard)")
    c.fill = FILL_SUBHDR
    c.font = Font(bold=False, color="FFFFFF", name="Calibri", size=10)
    c.alignment = CENTER

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    # ---- Column headers ----
    HDR_ROW = 4
    fixed_cols = ["#", "G-TRN Standard Role", "Abbr", "LoB", "Mandatory"]
    dyn_col_start = len(fixed_cols) + 1   # column index where project columns start

    for ci, h in enumerate(fixed_cols, start=1):
        c = apply_cell(ws, HDR_ROW, ci, h, fill=FILL_HEADER,
                       font=FONT_WHITE_BOLD, alignment=CENTER)

    for pi, proj in enumerate(projects):
        col = dyn_col_start + pi * 2 - 1
        apply_cell(ws, HDR_ROW, col,
                   f"{proj['name']}\nStatus",
                   fill=FILL_SUBHDR, font=FONT_WHITE_BOLD, alignment=CENTER)
        apply_cell(ws, HDR_ROW, col + 1,
                   f"{proj['name']}\nMatched Role / Note",
                   fill=FILL_SUBHDR, font=FONT_WHITE_BOLD, alignment=CENTER)

    ws.row_dimensions[HDR_ROW].height = 40

    # ---- Data rows ----
    prev_lob = None
    current_row = HDR_ROW + 1

    all_matched_titles = {proj["name"]: set() for proj in projects}

    for ri, std in enumerate(STANDARD_ROLES, start=1):
        lob = std["lob"]

        # Section separator
        if lob != prev_lob:
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row,
                end_column=fixed_cols.__len__() + len(projects) * 2
            )
            lob_label = {"SRS+VHC": "CORE ROLES (SRS + VHC)",
                         "SRS": "SRS-SPECIFIC ROLES",
                         "VHC": "VHC-SPECIFIC ROLES"}.get(lob, lob)
            c = ws.cell(row=current_row, column=1, value=f"  {lob_label}")
            c.fill = FILL_SUBHDR
            c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            c.alignment = LEFT
            c.border = THIN_BORDER
            ws.row_dimensions[current_row].height = 18
            current_row += 1
            prev_lob = lob

        row_fill = FILL_MAND if std["mandatory"] else None

        apply_cell(ws, current_row, 1, ri, fill=row_fill, alignment=CENTER)
        apply_cell(ws, current_row, 2, std["role"],
                   fill=row_fill,
                   font=FONT_BOLD if std["mandatory"] else FONT_NORMAL,
                   alignment=LEFT)
        apply_cell(ws, current_row, 3, std["abbr"],  fill=row_fill, alignment=CENTER)
        apply_cell(ws, current_row, 4, std["lob"],   fill=row_fill, alignment=CENTER)
        apply_cell(ws, current_row, 5,
                   "YES" if std["mandatory"] else "if applicable",
                   fill=row_fill, alignment=CENTER)

        for pi, proj in enumerate(projects):
            col = dyn_col_start + pi * 2 - 1
            result = match_role(std["role"], proj["roles"])
            sf = status_fill(result["status"])

            apply_cell(ws, current_row, col,
                       status_icon(result["status"]),
                       fill=sf, alignment=CENTER)
            apply_cell(ws, current_row, col + 1,
                       f"{result['matched_title']}\n({result['entity']})"
                       if result["matched_title"] else "— role not found —",
                       fill=sf, alignment=LEFT)

            if result["matched_title"]:
                all_matched_titles[proj["name"]].add(result["matched_title"])

        ws.row_dimensions[current_row].height = 30
        current_row += 1

    # ---- Extra roles section ----
    current_row += 1
    ws.merge_cells(
        start_row=current_row, start_column=1,
        end_row=current_row,
        end_column=fixed_cols.__len__() + len(projects) * 2
    )
    c = ws.cell(row=current_row, column=1,
                value="  EXTRA ROLES IN PROJECT OBS (no direct G-TRN standard equivalent — may be project-specific or legacy GTS roles)")
    c.fill = PatternFill("solid", fgColor="4472C4")
    c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    c.alignment = LEFT
    c.border = THIN_BORDER
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    # Collect all extras per project
    max_extras = 0
    all_extras = {}
    for proj in projects:
        extras = find_extra_roles(proj["roles"], all_matched_titles[proj["name"]])
        # Remove admin/support roles that are clearly not PTO
        extras = [e for e in extras if not any(
            skip in e["title"].lower()
            for skip in ["admin", "document control", "executive", "engineer",
                         "specialist", "architect", "officer", "lead", "po admin"]
        )]
        all_extras[proj["name"]] = extras
        max_extras = max(max_extras, len(extras))

    for i in range(max_extras):
        apply_cell(ws, current_row, 1, "", alignment=CENTER)
        apply_cell(ws, current_row, 2, f"Extra role #{i+1}", alignment=LEFT)
        apply_cell(ws, current_row, 3, "", alignment=CENTER)
        apply_cell(ws, current_row, 4, "", alignment=CENTER)
        apply_cell(ws, current_row, 5, "", alignment=CENTER)

        for pi, proj in enumerate(projects):
            col = dyn_col_start + pi * 2 - 1
            extras = all_extras[proj["name"]]
            if i < len(extras):
                e = extras[i]
                apply_cell(ws, current_row, col,
                           STATUS_EXTRA, fill=FILL_BLUE, alignment=CENTER)
                apply_cell(ws, current_row, col + 1,
                           f"{e['title']}\n({e['entity']})",
                           fill=FILL_BLUE, alignment=LEFT)
            else:
                apply_cell(ws, current_row, col, "", alignment=CENTER)
                apply_cell(ws, current_row, col + 1, "", alignment=LEFT)

        ws.row_dimensions[current_row].height = 28
        current_row += 1

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    for pi in range(len(projects)):
        col_s = get_column_letter(dyn_col_start + pi * 2 - 1)
        col_n = get_column_letter(dyn_col_start + pi * 2)
        ws.column_dimensions[col_s].width = 16
        ws.column_dimensions[col_n].width = 36

    ws.freeze_panes = f"F{HDR_ROW + 1}"


# ---------------------------------------------------------------------------
# 6.  SHEET 2 – PROJECT DETAIL (one sheet per project)
# ---------------------------------------------------------------------------
def build_project_detail(wb: Workbook, proj: dict) -> None:
    ws = wb.create_sheet(title=proj["name"][:31])
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1,
                value=f"OBS ROLE DETAIL  |  Project: {proj['name']}")
    c.fill = FILL_HEADER
    c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Sub-title
    ws.merge_cells("A2:G2")
    c = ws.cell(row=2, column=1,
                value=f"Source OBS: {proj.get('source', 'N/A')}  |  Entities: {proj.get('entities', 'N/A')}  |  Comparison Date: {proj.get('date', '')}")
    c.fill = FILL_SUBHDR
    c.font = Font(color="FFFFFF", name="Calibri", size=10)
    c.alignment = CENTER
    ws.row_dimensions[2].height = 18

    # Headers
    headers = ["#", "G-TRN Standard Role", "LoB", "Mandatory",
               "Status", "Matched OBS Role Title", "Entity in OBS"]
    for ci, h in enumerate(headers, 1):
        c = apply_cell(ws, 4, ci, h, fill=FILL_HEADER,
                       font=FONT_WHITE_BOLD, alignment=CENTER)
    ws.row_dimensions[4].height = 32

    prev_lob = None
    row = 5
    for ri, std in enumerate(STANDARD_ROLES, 1):
        lob = std["lob"]
        if lob != prev_lob:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=7)
            lob_label = {"SRS+VHC": "CORE ROLES (SRS + VHC)",
                         "SRS": "SRS-SPECIFIC ROLES",
                         "VHC": "VHC-SPECIFIC ROLES"}.get(lob, lob)
            c = ws.cell(row=row, column=1, value=f"  {lob_label}")
            c.fill = FILL_SUBHDR
            c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            c.alignment = LEFT
            c.border = THIN_BORDER
            ws.row_dimensions[row].height = 16
            row += 1
            prev_lob = lob

        result = match_role(std["role"], proj["roles"])
        sf = status_fill(result["status"])
        mf = FILL_MAND if std["mandatory"] else None
        row_fill = sf  # status colour takes priority

        apply_cell(ws, row, 1, ri, fill=mf, alignment=CENTER)
        apply_cell(ws, row, 2, std["role"],
                   fill=mf,
                   font=FONT_BOLD if std["mandatory"] else FONT_NORMAL,
                   alignment=LEFT)
        apply_cell(ws, row, 3, std["lob"], fill=mf, alignment=CENTER)
        apply_cell(ws, row, 4, "YES" if std["mandatory"] else "if appl.",
                   fill=mf, alignment=CENTER)
        apply_cell(ws, row, 5, status_icon(result["status"]),
                   fill=row_fill, alignment=CENTER)
        apply_cell(ws, row, 6,
                   result["matched_title"] if result["matched_title"] else "— NOT FOUND —",
                   fill=row_fill, alignment=LEFT)
        apply_cell(ws, row, 7, result["entity"] or "—",
                   fill=row_fill, alignment=LEFT)
        ws.row_dimensions[row].height = 22
        row += 1

    # ---- All roles in this OBS ----
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1, value="  ALL ROLES FOUND IN THIS PROJECT OBS")
    c.fill = PatternFill("solid", fgColor="4472C4")
    c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    c.alignment = LEFT
    c.border = THIN_BORDER
    ws.row_dimensions[row].height = 16
    row += 1

    headers2 = ["#", "OBS Role Title", "Entity", "", "", "", ""]
    for ci, h in enumerate(headers2, 1):
        apply_cell(ws, row, ci, h, fill=FILL_SUBHDR,
                   font=FONT_WHITE_BOLD, alignment=CENTER)
    ws.row_dimensions[row].height = 20
    row += 1

    seen_titles = set()
    counter = 1
    for r in proj["roles"]:
        key = (r["title"].lower(), r["entity"])
        if key in seen_titles:
            continue
        seen_titles.add(key)
        # Check if it matched anything
        matched_any = any(
            match_role(s["role"], proj["roles"])["matched_title"].lower() == r["title"].lower()
            for s in STANDARD_ROLES
        )
        rf = FILL_GREEN if matched_any else FILL_BLUE
        apply_cell(ws, row, 1, counter, fill=rf, alignment=CENTER)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        c = ws.cell(row=row, column=2, value=r["title"])
        c.fill = rf
        c.font = FONT_NORMAL
        c.alignment = LEFT
        c.border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        c2 = ws.cell(row=row, column=6, value=r["entity"])
        c2.fill = rf
        c2.font = FONT_ITALIC
        c2.alignment = LEFT
        c2.border = THIN_BORDER
        ws.row_dimensions[row].height = 18
        row += 1
        counter += 1

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 20
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# 7.  SHEET 3 – SUMMARY SCORECARD
# ---------------------------------------------------------------------------
def build_scorecard(wb: Workbook, projects: list[dict]) -> None:
    ws = wb.create_sheet(title="Scorecard")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    c = ws.cell(row=1, column=1, value="OBS COMPLIANCE SCORECARD  —  All Projects vs G-TRN A08001")
    c.fill = FILL_HEADER
    c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    c = ws.cell(row=2, column=1,
                value="Mandatory roles = 7  |  SRS-specific = 5  |  VHC-specific = 5  |  Total standard = 17")
    c.fill = FILL_SUBHDR
    c.font = Font(color="FFFFFF", name="Calibri", size=10)
    c.alignment = CENTER
    ws.row_dimensions[2].height = 16

    headers = ["Project", "Source", "Entities",
               "Matched", "Partial", "Missing",
               "% Mandatory\nCovered", "% Total\nCovered", "Risk Level"]
    for ci, h in enumerate(headers, 1):
        apply_cell(ws, 4, ci, h, fill=FILL_HEADER,
                   font=FONT_WHITE_BOLD, alignment=CENTER)
    ws.row_dimensions[4].height = 36

    mandatory_roles = [s for s in STANDARD_ROLES if s["mandatory"]]
    total_roles = len(STANDARD_ROLES)

    for pi, proj in enumerate(projects):
        row = 5 + pi
        counts = {STATUS_MATCHED: 0, STATUS_PARTIAL: 0, STATUS_MISSING: 0}
        mand_covered = 0

        for std in STANDARD_ROLES:
            r = match_role(std["role"], proj["roles"])
            counts[r["status"]] = counts.get(r["status"], 0) + 1
            if std["mandatory"] and r["status"] in (STATUS_MATCHED, STATUS_PARTIAL):
                mand_covered += 1

        matched_n = counts[STATUS_MATCHED]
        partial_n = counts[STATUS_PARTIAL]
        missing_n = counts[STATUS_MISSING]
        pct_mand  = round(mand_covered / len(mandatory_roles) * 100)
        pct_total = round((matched_n + partial_n) / total_roles * 100)

        risk = ("LOW" if pct_mand >= 85 else
                "MEDIUM" if pct_mand >= 60 else "HIGH")
        risk_fill = (FILL_GREEN if risk == "LOW" else
                     FILL_ORANGE if risk == "MEDIUM" else FILL_RED)

        apply_cell(ws, row, 1, proj["name"], fill=None, font=FONT_BOLD, alignment=LEFT)
        apply_cell(ws, row, 2, proj.get("source", ""), alignment=LEFT)
        apply_cell(ws, row, 3, proj.get("entities", ""), alignment=CENTER)
        apply_cell(ws, row, 4, matched_n, fill=FILL_GREEN, alignment=CENTER)
        apply_cell(ws, row, 5, partial_n, fill=FILL_ORANGE, alignment=CENTER)
        apply_cell(ws, row, 6, missing_n, fill=FILL_RED, alignment=CENTER)
        apply_cell(ws, row, 7, f"{pct_mand}%",
                   fill=FILL_GREEN if pct_mand >= 85 else
                        FILL_ORANGE if pct_mand >= 60 else FILL_RED,
                   alignment=CENTER, font=FONT_BOLD)
        apply_cell(ws, row, 8, f"{pct_total}%",
                   fill=FILL_GREEN if pct_total >= 75 else
                        FILL_ORANGE if pct_total >= 50 else FILL_RED,
                   alignment=CENTER, font=FONT_BOLD)
        apply_cell(ws, row, 9, risk, fill=risk_fill,
                   font=FONT_BOLD, alignment=CENTER)
        ws.row_dimensions[row].height = 22

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["F"].width = 11
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 13
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# 8.  SHEET 4 – LEGEND & INSTRUCTIONS
# ---------------------------------------------------------------------------
def build_legend(wb: Workbook) -> None:
    ws = wb.create_sheet(title="Legend & How To Use")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    c = ws.cell(row=1, column=1,
                value="LEGEND & HOW TO USE THIS WORKBOOK")
    c.fill = FILL_HEADER
    c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 30

    rows = [
        (2, "COLOUR CODING", FILL_SUBHDR, FONT_WHITE_BOLD),
        (3, None, None, None),
        (4, "✔ Matched (Green)",   FILL_GREEN,  FONT_BOLD),
        (5, "~ Partial (Orange)",  FILL_ORANGE, FONT_BOLD),
        (6, "✘ Missing (Red)",     FILL_RED,    FONT_BOLD),
        (7, "+ Extra (Blue)",      FILL_BLUE,   FONT_BOLD),
        (8, "Yellow row",          FILL_MAND,   FONT_BOLD),
        (9, None, None, None),
        (10, "DESCRIPTIONS", FILL_SUBHDR, FONT_WHITE_BOLD),
        (11, None, None, None),
    ]

    descs = {
        4: "The existing project OBS has a role that directly matches the G-TRN standard role (keyword or name match ≥ 90%).",
        5: "A role with a similar title exists but uses a different name or is at a different level. Action: review and align naming.",
        6: "The standard G-TRN role has NO equivalent in the existing project OBS. This is a GAP that must be addressed.",
        7: "A role exists in the project OBS but has no equivalent in the G-TRN standard. May be legacy GTS roles or project-specific.",
        8: "Yellow background rows = MANDATORY core roles required for ALL projects (SRS and VHC). These must be staffed.",
    }

    for r_num, label, fill, font in rows:
        if label is None:
            ws.row_dimensions[r_num].height = 6
            continue
        ws.merge_cells(start_row=r_num, start_column=1,
                       end_row=r_num, end_column=2)
        c = ws.cell(row=r_num, column=1, value=f"  {label}")
        c.fill = fill or PatternFill()
        c.font = font or FONT_NORMAL
        c.alignment = LEFT
        c.border = THIN_BORDER
        if r_num in descs:
            ws.merge_cells(start_row=r_num, start_column=3,
                           end_row=r_num, end_column=5)
            d = ws.cell(row=r_num, column=3, value=descs[r_num])
            d.fill = fill or PatternFill()
            d.font = FONT_NORMAL
            d.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            d.border = THIN_BORDER
        ws.row_dimensions[r_num].height = 32

    row = 12
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value="  HOW TO ADD MORE PROJECTS")
    c.fill = FILL_SUBHDR
    c.font = FONT_WHITE_BOLD
    c.alignment = LEFT
    c.border = THIN_BORDER
    ws.row_dimensions[row].height = 20
    row += 1

    instructions = [
        "1. Open obs_compare.py in VS Code.",
        "2. In the PROJECTS list near the bottom of the file, add a new entry following the same pattern.",
        "3. For each new project supply: name, source (PDF or VSD filename), entities, date, and the roles list.",
        "   Roles can be added manually (as dicts with 'title' and 'entity') or auto-extracted from a PDF.",
        "4. Run:  python obs_compare.py",
        "5. A new sheet tab and two new columns will be added automatically to the Gap Analysis sheet.",
        "",
        "TIPS:",
        "- The matching uses keyword + fuzzy matching, so role titles don't need to be exact.",
        "- For Visio (.vsd) files, export them to PDF first, then place in 1_Input/.",
        "- The Scorecard sheet gives you a quick at-a-glance compliance % for each project.",
    ]
    for inst in instructions:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=inst)
        c.font = FONT_NORMAL if inst and not inst.startswith("TIP") else FONT_BOLD
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
        ws.row_dimensions[row].height = 18
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20


# ---------------------------------------------------------------------------
# 9.  MAIN
# ---------------------------------------------------------------------------
def main():
    output_dir = r"\\spiderman\DEPARTMENTS\Project_Management_Office\Artificial_Intelligence\26_OBS Compare\2_Output"
    os.makedirs(output_dir, exist_ok=True)

    # ---- Define projects to compare ----
    # Each project has: name, source, entities, date, roles (list of dicts)
    # To add a new project: duplicate the dict block and change the data.
    projects = [
        {
            "name":     "R152B SA1 URS",
            "source":   "R152B SA1 URS_Project OBSFinal (1).pdf",
            "entities": "HR Singapore | HR France | HR Canada",
            "date":     "2026-06-17",
            "roles":    manual_r152b_roles(),
        },
        # ---- ADD FUTURE PROJECTS HERE ----
        # {
        #     "name":     "ProjectXYZ",
        #     "source":   "1_Input/ProjectXYZ_OBS.pdf",
        #     "entities": "HR Italy | HR UK",
        #     "date":     "2026-06-17",
        #     "roles":    extract_obs_roles_from_pdf(
        #                     r"\\spiderman\DEPARTMENTS\...\ProjectXYZ_OBS.pdf"),
        # },
    ]

    print("Building OBS Gap Analysis workbook ...")

    wb = Workbook()

    build_gap_matrix(wb, projects)
    for proj in projects:
        build_project_detail(wb, proj)
    build_scorecard(wb, projects)
    build_legend(wb)

    # Put Scorecard as second sheet for easy access
    wb.move_sheet("Scorecard", offset=1)

    out_path = os.path.join(output_dir, "OBS_Gap_Analysis.xlsx")
    wb.save(out_path)
    print(f"\n✔ Saved: {out_path}")
    print("\nSummary:")
    for proj in projects:
        m = p = miss = 0
        for std in STANDARD_ROLES:
            r = match_role(std["role"], proj["roles"])
            if r["status"] == STATUS_MATCHED:   m += 1
            elif r["status"] == STATUS_PARTIAL:  p += 1
            else:                                miss += 1
        print(f"  {proj['name']:30s}  Matched={m}  Partial={p}  MISSING={miss}")


if __name__ == "__main__":
    main()
